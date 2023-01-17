#Important to tell Tammy
#Must change names on CST to match file names given.

from ast import While
from openpyxl import load_workbook
import PySimpleGUI as sg



layout = [  [sg.LBox([], size=(20,10), key='-FILESLB-')],
            [sg.Input(visible=False, enable_events=True, key='-IN-'), sg.FilesBrowse()],
            [sg.Button('Go'), sg.Button('Exit')]  ]

window = sg.Window('Window Title', layout)

while True:             # Event Loop
    event, values = window.read()
    if event in (sg.WIN_CLOSED, 'Exit'):
        break
    # When choice has been made, then fill in the listbox with the choices
    if event == '-IN-':
        window['-FILESLB-'].Update(values['-IN-'].split(';'))
window.close()

#Obtain name of worksheet to be added
wb2Name = ('LGLSCBS  0223-309F  OL.xlsx')

#prepare Cohort Sched excel workbook and sheet to be added
wb1 = load_workbook('Cohort Schedules Tracking test.xlsx')
wb2 = load_workbook(wb2Name)

#Obtain primary worksheet from source
wsC = wb2['Primary']

#Split name of wb2 to isolate name of program
targetWS = wb2Name.split()
ws = wb1[targetWS[0]]


#Find number of rows to be copied and inserted. compare val of a and b cells, 
#increase b by one (one row down) and repeat until section is over. a starts at 3 to make up for header
#b starts at i (4) right under 3, I is incremented each comparison.
#The section name is used to find the number of courses.
i = 4 
a = ws.cell(row=3, column=1)
b = ws.cell(row=i, column=1)
counter = 2

while (a.value == b.value):
    counter += 1
    i += 1
    b = ws.cell(row=i, column=1)

ws.insert_rows(2, counter)

#copy header of section to obtain style/color
ws.cell(row=2, column=1).style = ws.cell(row=(counter+2), column=1).style
ws.cell(row=2, column=1).value = targetWS[0] + " " + targetWS[1]

#obtain section from copied ws and paste under 'program'
for row in ws.iter_rows(min_row=3, max_col=1, max_row=counter+1):
    for cell in row:
        cell.value = targetWS[1]

#obtain day and paste under Day section
for row in ws.iter_rows(min_row=3, min_col=4, max_col=4, max_row=counter+1):
    for cell in row:
        cell.value = targetWS[2].replace('.xlsx', '')

#Find cells with a 'credit' value, meaning rows with course listing on them. Obtain global 'term' for current term as wel
term = 'none'
def iterate():
    for row in wsC.iter_rows():
        for cell in row:
            if cell in cellChecker: #make sure not to repeat the same cell
                continue

            if cell.value is not None:
                a = str(cell.value).split()
                if (a[0] == 'Term'):
                    global term
                    cellChecker.append(cell)
                    term = a[2]
                #Find a course row based on a number existing on column B (the credits cell)
                if cell.column_letter == 'B' and str(cell.value).isnumeric():
                    cellChecker.append(cell)
                    return cell
    
rowCounter=0
row = 3
col = 2
currentRow = 0
currentCol = 0
global cellChecker
cellChecker = []

#row, col determine coordinates in CST doc. currentRow currentCol determine cooridnates on schedule to be added.
#Counter is subtracted 1 to account for op header rows 1 and 2 
while(rowCounter != counter-1):
    print("Loading...", end="\r")
    #Find cells with a 'credit' value, meaning rows with course listing on them. Obtain global 'term' for current term as well.
    currentCell = iterate()
    
    #Write down course name and number
    ws.cell(row, col).value = str(currentCell.offset(0, -1).value).split()[0] + " " + str(currentCell.offset(0, -1).value).split()[1].replace(':','')
    #Write down term
    ws.cell(row, col+1).value = term
    #Write down dates
    ws.cell(row, col+4).value = str(currentCell.offset(0, 1).value).split()[0]
    ws.cell(row, col+5).value = str(currentCell.offset(0, 1).value).split()[2]

    #Move to next row for next iteration
    row += 1
    rowCounter += 1

wb1.save('Cohort Schedules Tracking test2.xlsx')
print('I am finished.')