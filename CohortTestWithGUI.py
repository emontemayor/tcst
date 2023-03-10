#MIT License

#Copyright (c) 2023 Emiliano Montemayor

#Permission is hereby granted, free of charge, to any person obtaining a copy
#of this software and associated documentation files (the "Software"), to deal
#in the Software without restriction, including without limitation the rights
#to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
#copies of the Software, and to permit persons to whom the Software is
#furnished to do so, subject to the following conditions:

#The above copyright notice and this permission notice shall be included in all
#copies or substantial portions of the Software.

#THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
#IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
#FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
#AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
#LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
#OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
#SOFTWARE.

# Python program to create
# a file explorer in Tkinter
# import filedialog module
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from tkinter import filedialog
import tkinter as tk
import os

class App:
    def __init__(self, root):

#/--------------------------------GUI------------------------------/
        #setting title
        root.title("Tammy's Cohort Sched Tool")
        #setting window size
        root.geometry("450x180")
        root.resizable(width=False, height=False)

        # Source file label
        source_label = tk.Label(root, text="Calendar:")
        source_label.place(x=30,y=30)

        self.source_file_label = tk.Label(root, text='No File Selected')
        self.source_file_label.place(x=150,y=30)

        # Source file browse button
        source_browse_button = tk.Button(root, text="Browse", command=self.browse_source, bg="#8AC6D1", fg="white", width=10, height=1, relief='raised')
        source_browse_button.place(x=320,y=30)
        
        # Target file label
        target_label = tk.Label(root, text="Cohort Schedule:")
        target_label.place(x=30,y=70)

        self.target_file_label = tk.Label(root, text='No File Selected')
        self.target_file_label.place(x=150,y=70)

        # Target file browse button
        target_browse_button = tk.Button(root, text="Browse", command=self.browse_target, bg="#8AC6D1", fg="white", width=10, height=1, relief='raised')
        target_browse_button.place(x=318,y=70)

        # Submit button
        submit_button = tk.Button(root, text="Submit", command=self.submit, bg="#8AC6D1", fg="white", width=20, height=2)
        submit_button.place(x=150,y=115)

        #Status notification
        self.status_label = tk.Label(root, text="")
        self.status_label.place(x=330, y=125)

    def browse_source(self):
        source_file = filedialog.askopenfilename(initialdir = "/", title = "Select a File", filetypes = (("Excel files", "*.xlsx"),("Excel files", "*.xls")))
        self.source_file_label.config(text=os.path.basename(source_file))
        self.status_label.config(text="")

    def browse_target(self):
        target_file = filedialog.askopenfilename(initialdir = "/", title = "Select a File", filetypes = (("Excel files", "*.xlsx"),("Excel files", "*.xls")))
        self.target_file_label.config(text=os.path.basename(target_file))
        self.status_label.config(text="")

#/--------------------------------Submit Algorithm------------------------------/
    #Submit button function. Will attempt to load CST.xlsx file first, then impor the data
    #from the calendar into the CST.xlsx file.
    def submit(self):
        try:
            wb1.save(target_file_name)
            self.status_label.config(text="Loading..")
        except: 
            self.status_label.config(text="Error: Close Excel Files")

        # here you have access to the source and target files
        source_file_name = self.source_file_label['text']
        target_file_name = self.target_file_label['text']

        #prepare Cohort Sched excel workbook and sheet to be added
        wb1 = load_workbook(target_file_name)
        wb2 = load_workbook(source_file_name)
        
        date_style = NamedStyle(name='date_style')
        date_style.number_format='DD/MM/YYYY'

        number_style = NamedStyle(name='number_style')
        number_style.number_format='0.00E+00'

        #Obtain primary worksheet from source. Check if Template is used instead of primary. Otherwise ask user to switch WS names.
        try: 
            wsC = wb2['Primary']
        except:
            try:
                wsC = wb2['Template']
            except: 
                self.status_label.config(text="Workbook Name Mismatch")

        #Split name of wb2 to isolate name of program
        targetWS = source_file_name.split()
        ws = wb1[targetWS[0]]

        #Find number of rows to be copied and inserted. compare val of a and b cells, 
        #increase b by one (one row down) and repeat until section is over. a starts at 3 to make up for header
        #b starts at i (4) right under 3, I is incremented each comparison.
        #The section name is used to find the number of courses.
        i = 4 
        a = ws.cell(row=3, column=1)
        b = ws.cell(row=i, column=1)
        counter = 2

        while (a.value == b.value):
            counter += 1
            i += 1
            b = ws.cell(row=i, column=1)

        ws.insert_rows(2, counter)
        ws.cell(row=2, column=1).value = targetWS[0] + " " + targetWS[1]

        #obtain section from copied ws and paste under 'program'
        for row in ws.iter_rows(min_row=3, max_col=1, max_row=counter+1):
            for cell in row:
                cell.value = targetWS[1]

        #obtain day and paste under Day section
        for row in ws.iter_rows(min_row=3, min_col=4, max_col=4, max_row=counter+1):
            for cell in row:
                cell.value = targetWS[2]
                if ".xlsx" in cell.value:
                    cell.value = cell.value.replace(".xlsx", "") 

        #Find cells with a 'credit' value, meaning rows with course listing on them. Obtain global 'term' for current term as wel
        def iterate():
            for row in wsC.iter_rows():
                for cell in row:
                    if cell in cellChecker: #make sure not to repeat the same cell
                        continue

                    if cell.value is not None:
                        a = str(cell.value).split()
                        if (a[0] == 'Term'):
                            global term
                            cellChecker.append(cell)
                            term = a[2]
                        #Find a course row based on a number existing on column B (the credits cell)
                        if cell.column_letter == 'B' and str(cell.value).isnumeric():
                            cellChecker.append(cell)
                            return cell
    
        rowCounter=0
        row = 3
        col = 2
        currentRow = 0
        currentCol = 0
        global cellChecker
        cellChecker = []

        #row, col determine coordinates in CST doc. currentRow currentCol determine cooridnates on schedule to be added.
        #Counter is subtracted 1 to account for op header rows 1 and 2 
        while(rowCounter != counter-1):
            #Find cells with a 'credit' value, meaning rows with course listing on them. Obtain global 'term' for current term as well.
            currentCell = iterate()
    
            #Write down course name and number
            ws.cell(row, col).value = str(currentCell.offset(0, -1).value).split()[0] + " " + str(currentCell.offset(0, -1).value).split()[1].replace(':','')

            #Write down term
            ws.cell(row, col+1).style = 'number_style'
            ws.cell(row, col+1).value = term

            #write down weeks
            ws.cell(row, col+3).style = 'number_style'
            ws.cell(row, col+3).value = str(currentCell.offset(0, 2).value)

            #Write down dates
            ws.cell(row, col+5).style = 'date_style'
            ws.cell(row, col+6).style = 'date_style'
            ws.cell(row, col+5).value = str(currentCell.offset(0, 1).value).split()[0]
            ws.cell(row, col+6).value = str(currentCell.offset(0, 1).value).split()[2]         

            #Move to next row for next iteration
            row += 1
            rowCounter += 1

        #show complete flag and save.
        try:
            wb1.save(target_file_name)
            self.status_label.config(text="Complete!")
        except: 
            self.status_label.config(text="Error: Close Excel Files")

root = tk.Tk()
app = App(root)
root.mainloop()


